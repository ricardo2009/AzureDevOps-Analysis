from azure.devops.connection import Connection
from msrest.authentication import BasicAuthentication
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Autenticação
personal_access_token = ' '
organization_url = 'https://dev.azure.com/nome_da_empresa'
credentials = BasicAuthentication('', personal_access_token)
connection = Connection(base_url=organization_url, creds=credentials)

# Clientes
core_client = connection.clients.get_core_client()
git_client = connection.clients.get_git_client()
build_client = connection.clients.get_build_client()

# Dados consolidados
workbook = Workbook()

for project in core_client.get_projects():
  project_name = project.name
  print(f"Processando projeto: {project_name}")

  # Repositórios
  repositories = git_client.get_repositories(project_name)
  repo_data = []
  for repository in repositories:
    repo_data.append({
      "Name": repository.name,
      "ID": repository.id,
      "Default Branch": repository.default_branch,
      "Is Disabled": repository.is_disabled,
      "Size": repository.size,
      "Remote URL": repository.remote_url,
      "SSH URL": repository.ssh_url,
      "URL": repository.url,
      "Web URL": repository.web_url
    })
  print(f"Encontrados {len(repo_data)} repositórios")

  # Pipelines
  pipelines = build_client.get_builds(project_name)
  pipeline_data = []
  for pipeline in pipelines:
    pipeline_data.append({
      "ID": pipeline.id,
      "Name": pipeline.definition.name,
      "Status": pipeline.result,
      "URL": pipeline.url
    })
  print(f"Encontradas {len(pipeline_data)} pipelines")

  # Commits
  commits_data = []
  for repository in repositories:
    try:
      commits = git_client.get_commits(repository_id=repository.id, project=project_name, search_criteria=None)
      for commit in commits:
        commits_data.append({
          "Date": commit.committer.date.replace(tzinfo=None),
          "Email": commit.committer.email,
          "Name": commit.committer.name,
          "Comment": commit.comment,
          "URL": commit.url
        })
    except Exception as e:
      print(e)
  print(f"Encontrados {len(commits_data)} commits")

# Get information agents
agent_client = connection.clients_v7_0.get_task_agent_client()

# Create a DataFrame to store the agent data
agent_data = []
for project in core_client.get_projects():
  project_name = project.name
  print(f"Processando projeto: {project_name}")
  
  agent_pools = agent_client.get_agent_pools()
  for pool in agent_pools:
    pool_id = pool.id
    pool_name = pool.name

    # Get the agents for the pool
    agents = agent_client.get_agents(pool_id)

    for agent in agents:
      # Get the projects that the agent can serve
      agent_projects = []
      try:
        agent_client.get_agent(pool_id, agent.id, project_name)
        agent_projects.append(project_name)
      except:
        pass

      # Add the agent data to the DataFrame
      agent_data.append({
        "Pool Name": pool_name,
        "Agent Name": agent.name,
        "Status": agent.status,
        "Version": agent.version,
        "Projects": ", ".join(agent_projects)
      })

  print(f"Encontrados {len(agent_data)} agents")

# Adicionar dados à planilha
agent_sheet = workbook.create_sheet(title=f"Agents - {project_name}")
agent_sheet.append([])
agent_sheet.append(["Agents"])
for row in dataframe_to_rows(pd.DataFrame(agent_data), index=False, header=True):
  agent_sheet.append(row)

  # Criar planilha com os dados
  df_repositories = pd.DataFrame(repo_data)
  df_pipelines = pd.DataFrame(pipeline_data)
  df_commits = pd.DataFrame(commits_data)

  # Adicionar dados à planilha
  repositories_sheet = workbook.create_sheet(title=f"Data - {project_name}")

  repositories_sheet.append([])
  repositories_sheet.append(["Repositories"])
  for row in dataframe_to_rows(df_repositories, index=False, header=True):
    repositories_sheet.append(row)
  repositories_sheet.append([])
  repositories_sheet.append(["Pipelines"])
  for row in dataframe_to_rows(df_pipelines, index=False, header=True):
    repositories_sheet.append(row)
  repositories_sheet.append([])
  repositories_sheet.append(["Commits"])
  for row in dataframe_to_rows(df_commits, index=False, header=True):
    repositories_sheet.append(row)

# Salvar o arquivo Excel uma única vez
workbook.save("analysis_result.xlsx")
print("Arquivo Excel salvo com sucesso!")

